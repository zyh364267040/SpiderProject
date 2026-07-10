# -*- coding: utf-8 -*-
"""美国人口普查局 Census ZIP Code 就业与收入数据采集脚本。

根据 ZIP Code 采集 ACS 就业、职业类型和家庭收入区间等公开统计数据，并写入 Excel。
历史 API Key 已删除。
"""
# @Time: 2022/9/26 10:57
import os.path
import ast
import sys
import time

import openpyxl
import requests
import random
import pandas as pd


user_agents = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.79 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.53 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:101.0) Gecko/20100101 Firefox/101.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A',
]


def get_data(url, params=None):
    headers = {
        'User-Agent': random.choice(user_agents)
    }
    for i in range(5):
        try:
            print('正在请求:', url)
            res = requests.get(url, headers=headers, params=params)
            # print(res.text)
            # print(res.url)
            time.sleep(random.choice([2, 3, 4]))
        except Exception as e:
            print(e)
        else:
            break

    return res


def main():
    # 新建excel
    if os.path.exists('census.xlsx'):
        wb = openpyxl.load_workbook('census.xlsx')
        sheet = wb.active
        row = sheet.max_row
        i = row + 1
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'LONG_ZIP_CD'
        sheet['B1'] = 'Employed'
        sheet['C1'] = 'Unemployed'
        sheet['D1'] = 'Workers 16 years and over'
        sheet['E1'] = 'Worked from home'
        sheet['F1'] = 'Civilian employed population 16 years and over'
        sheet['G1'] = 'Private wage and salary workers'
        sheet['H1'] = 'Government workers'
        sheet['I1'] = 'Self-employed in own not incorporated business workers'
        sheet['J1'] = 'Unpaid family workers'
        sheet['K1'] = 'Less than $10,000'
        sheet['L1'] = '$10,000 to $14,999'
        sheet['M1'] = '$15,000 to $24,999'
        sheet['N1'] = '$25,000 to $34,999'
        sheet['O1'] = '$35,000 to $49,999'
        sheet['P1'] = '$50,000 to $74,999'
        sheet['Q1'] = '$75,000 to $99,999'
        sheet['R1'] = '$100,000 to $149,999'
        sheet['S1'] = '$150,000 to $199,999'
        sheet['T1'] = '$200,000 or more'
        sheet['U1'] = 'Median household income (dollars)'
        sheet['V1'] = 'Mean household income (dollars)'
        i = 2

    # 从excel中读取需要下载数据的zipcode
    # df = pd.read_excel('zipcode.xlsx', sheet_name='Sheet1')
    # zip_code_list = df['LONG_ZIP_CD'].to_list()
    # print(f'需要下载的zipcode:{zip_code_list}')

    with open('zipcode.txt', 'r', encoding='utf-8') as f:
        zip_code_list = ast.literal_eval(f.read())
    print(f'需要下载的zipcode:{zip_code_list}')


    # 获取地区编号
    area_url = 'https://api.census.gov/data/2020/acs/acs5?get=NAME&for=state:*'
    area_res = get_data(area_url)
    area_dic = area_res.json()

    # 获取每个地区的zipcode
    for area_list in area_dic[1:]:
        area = area_list[1]
        num_url = f'https://www.census.gov/acs/www/data/data-tables-and-tools/narrative-profiles/zcta2020/zcta{area}.json'
        num_res = get_data(num_url)
        num_dic = num_res.json()
        for num_list in num_dic[1:]:
            # https://data.census.gov/cedsci/table?tid=ACSDP5Y2020.DP03&g=0400000US42_860XX00US15005
            num = num_list[0]
            zip_code_num = num[-5:]
            if int(zip_code_num) in zip_code_list:
                print('开始获取zip_code:', zip_code_num)
                detail_url = 'https://data.census.gov/api/access/data/table'
                res_url = f'https://data.census.gov/cedsci/table?tid=ACSDP5Y2020.DP03&g=0400000US{area}_{num}'
                print(res_url)
                params = {
                    'g': f'{num}:0400000US{area}',
                    'id': 'ACSDP5Y2020.DP03'
                }
                detail_res = get_data(detail_url, params)
                detail_res_json = detail_res.json()
                data = detail_res_json['response']['data'][2]
                employed = data[93]
                print('employed', employed)
                unemployed = data[776]
                print('unemployed', unemployed)
                work_16_years_and_over = data[179]
                print('work_16_years_and_over', work_16_years_and_over)
                worked_from_home = data[283]
                print('worked_from_home', worked_from_home)
                civilian = data[685]
                print('civilian', civilian)
                private = data[339]
                print('private', private)
                government = data[284]
                print('government', government)
                self_employed = data[301]
                print('self_employed', self_employed)
                unpaid = data[1026]
                print('unpaid', unpaid)
                less_than = data[992]
                print('less_than', less_than)
                d10000 = data[1006]
                print('d10000', d10000)
                d15000 = data[949]
                print('d15000', d15000)
                d25000 = data[973]
                print('d25000', d25000)
                d35000 = data[915]
                print('d35000', d35000)
                d50000 = data[931]
                print('d50000', d50000)
                d75000 = data[873]
                print('d75000', d75000)
                d100000 = data[892]
                print('d100000', d100000)
                d150000 = data[509]
                print('d150000', d150000)
                d200000 = data[526]
                print('d200000', d200000)
                median = data[478]
                print('median', median)
                mean = data[497]
                print('mean', mean)

                # 把数据写入到excel
                sheet[f'A{i}'] = zip_code_num
                sheet[f'B{i}'] = employed
                sheet[f'C{i}'] = unemployed
                sheet[f'D{i}'] = work_16_years_and_over
                sheet[f'E{i}'] = worked_from_home
                sheet[f'F{i}'] = civilian
                sheet[f'G{i}'] = private
                sheet[f'H{i}'] = government
                sheet[f'I{i}'] = self_employed
                sheet[f'J{i}'] = unpaid
                sheet[f'K{i}'] = less_than
                sheet[f'L{i}'] = d10000
                sheet[f'M{i}'] = d15000
                sheet[f'N{i}'] = d25000
                sheet[f'O{i}'] = d35000
                sheet[f'P{i}'] = d50000
                sheet[f'Q{i}'] = d75000
                sheet[f'R{i}'] = d100000
                sheet[f'S{i}'] = d150000
                sheet[f'T{i}'] = d200000
                sheet[f'U{i}'] = median
                sheet[f'V{i}'] = mean
                i += 1

                index = zip_code_list.index(int(zip_code_num))
                zip_code_list.pop(index)
                print(zip_code_list)
                print(f'还有{len(zip_code_list)}个...')
                if not zip_code_list:
                    print('所有数据获取完成,终止程序!')
                    sys.exit()

    if len(zip_code_list) > 0:
        with open('zipcode.txt', 'w', encoding='utf-8') as f:
            f.write(str(zip_code_list))

    wb.save('census.xlsx')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('发生错误:', e)
