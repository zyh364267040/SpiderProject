# -*- coding = utf-8 -*-
# @Time: 2022/10/15 11:01
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor
import os
import sys
from selenium import webdriver


def get_data(url, i, row, xinghao_list):
    print(f'开始第{i}个,共有{row}个.')
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36',
    }
    for i in range(10):
        try:
            res = requests.get(url, headers=headers)
            # print(res.text)
            # time.sleep(random.choice([2, 3, 4]))
            print(f'请求成功:{url}')
            if res and res.status_code == 200:
                return res
            else:
                driver = webdriver.Chrome('/Users/zhouyanhui/Library/CloudStorage/OneDrive-个人/chromedriver')
                driver.get(url)
                k = input('出现错误,请查看原因,按任意键回车:')
                continue
        except Exception as e:
            print(url)
            print('出错了,正在重试...', e)
            if i == 9:
                with open('shibai.txt', 'a', encoding='utf-8') as f:
                    f.write(url + '\n')


# '"offers":[]}</script> <ti-tooltip slot="qrp-tooltip"> '
# def parse_data(res):
#     pattern = r'"offers":(.*?)}</script> <ti-tooltip slot="qrp-tooltip">'
#     result = re.findall(pattern, res.text)[0]
#     print(result)
#     result_list = json.loads(result)
#     for result in result_list:
#         sku = result['itemOffered']['sku']
#
#         print(sku)


def parse_data(res, name, xinghao_list):
    try:
        res_dic = res.json()
    except Exception as e:
        print('解析出错:', e)
        with open('name.txt', 'w', encoding='utf-8') as f:
            f.write(name + '\n')
        return None

    for k, v in res_dic.items():
        with open('ti.csv', 'a', encoding='utf-8') as f:
            print(f"{k}, {v['inventory']}")
            f.write(f"{k}, {v['inventory']}\n")
            xinghao_list.remove(name)
    print(f'{name}保存成功!')


def get_all_url(name, xinghao_list, i, row):
    # url = f'https://www.ti.com.cn/product/cn/{name}'
    url = f'https://www.ti.com.cn/productmodel/gpn/{name}/tistoresegmented'
    print(url)
    # 1. 发送请求获取响应
    res = get_data(url, i, row, xinghao_list)

    # 2. 解析数据
    parse_data(res, name, xinghao_list)


def save_data(data_list):
    # 新建excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = '型号'
    sheet['B1'] = '数量'
    i = 2

    for data in data_list:
        for k, v in data.items():
            sheet[f'A{i}'] = k
            sheet[f'B{i}'] = v
            i += 1

    wb.save('ti.xlsx')


def main():
    # 1. 读取excel
    wb = openpyxl.load_workbook('products.xlsx')
    sheet = wb.active
    row = sheet.max_row

    # 2. 保存数据的列表
    xinghao_list = []

    with ThreadPoolExecutor(16) as t:
        # 3. 逐条取出每一个
        for i in range(1, row+1):
            # print(f'开始第{i}个,共有{row}个.')
            name = sheet[f'A{i}'].value
            xinghao_list.append(name)

            # 4. 逐个发送请求
            t.submit(get_all_url, name, xinghao_list, i, row)
            # get_all_url(name, data_list)

    # print(data_list)
    # 5. 保存数据
    # save_data(data_list)
    with open('xinghao.txt', 'w', encoding='utf-8') as f:
        for xinghao in xinghao_list:
            f.write(xinghao + '\n')


if __name__ == '__main__':
    main()
