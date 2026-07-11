# -*- coding = utf-8 -*-
# @Time: 2022/10/14 13:51
import sys
import requests
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import re
import json
import random


# '"offers":[]}</script> <ti-tooltip slot="qrp-tooltip"> '
def parse_data(res, data_list, name):
    try:
        pattern = r'"offers":(.*?)}</script> <ti-tooltip slot="qrp-tooltip">'
        result = re.findall(pattern, res.text)[0]
        # print(result)
        result_list = json.loads(result)
        with open('name.csv', 'a', encoding='utf-8') as f:
            for result in result_list:
                sku = result['itemOffered']['sku']
                f.write(f'{sku},\n')
                print(sku)
            print(f'{name}获取完成!')
            data_list.remove(name)
    except Exception as e:
        print('发生错误:', e)


def get_data(url, data_list, num, row):
    headers = {
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36',
    }
    print(f'开始第{num}个,共有{row}个.')
    for i in range(10):
        try:
            res = requests.get(url, headers=headers, timeout=10)
            time.sleep(random.choice([2, 3, 4]))
            print(f'请求成功:{url}')
            if res and res.status_code == 200:
                return res
            else:
                with open('shibai.txt', 'w', encoding='utf-8') as f:
                    for name in data_list:
                        f.write(name + '\n')
                sys.exit()
        except Exception as e:
            print(url)
            print('出错了,正在重试...', e)
            if i == 9:
                with open('shibai.txt', 'w', encoding='utf-8') as f:
                    for name in data_list:
                        f.write(name + '\n')


# def parse_data(res, data_list, name):
#     try:
#         res_dic = res.json()
#         for k, v in res_dic.items():
#             with open('ti.csv', 'a', encoding='utf-8') as f:
#                 print(f"{k}, {v['inventory']}")
#                 f.write(f"{k}, {v['inventory']}\n")
#                 data_list.remove(name)
#         print(f'{name}保存成功!')
#     except Exception as e:
#         print('解析数据,发生错误:', e)


def get_all_url(name, data_list, num, row):
    url = f'https://www.ti.com.cn/product/cn/{name}'
    # url = f'https://www.ti.com.cn/productmodel/gpn/{name}/tistoresegmented'
    print(url)
    # 1. 发送请求获取响应
    res = get_data(url, data_list, num, row)
    # print(res.text)

    # 2. 解析数据
    parse_data(res, data_list, name)


def save_data(data_list):
    # 新建excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = '型号'
    sheet['B1'] = '数量'
    i = 2

    for data in data_list:
        for k, v in data.items():
            sheet[f'A{i}'] = k
            sheet[f'B{i}'] = v
            i += 1

    wb.save('ti.xlsx')


def main():
    # 1. 读取excel
    wb = openpyxl.load_workbook('products.xlsx')
    sheet = wb.active
    row = sheet.max_row

    name_list = []
    with open('name.csv', 'r', encoding='utf-8') as f:
        for name in f.readlines():
            name = name.strip()
            name_list.append(name)

    name_str = ''.join(name_list)

    # 2. 保存数据的列表
    data_list = []

    # 3. 逐条取出每一个
    # for i in range(1, row+1):
    #     name = sheet[f'A{i}'].value
    with open('shibai.txt', 'r', encoding='utf-8') as f:
        for name in f.readlines():
            name = name.replace('\n', '')
            if name in name_str:
                continue
            data_list.append(name)

    row = len(data_list)
    # print(data_list)
    # print(row)

    # 4. 逐个发送请求
    for num, name in enumerate(data_list):
        if name == 'TDA2EG-17' and name == 'TDA2EG-17' and name == 'CC1151-Q1':
            continue
        get_all_url(name, data_list, num+1, row)

    # print(data_list)
    # 5. 保存数据
    # save_data(data_list)


if __name__ == '__main__':
    main()
    # get_cookie()
