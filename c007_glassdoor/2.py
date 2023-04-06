# -*- coding = utf-8 -*-
# @Time: 2022/9/30 21:40
import requests
import os
import re
import json
import openpyxl
import random
import time
from lxml import etree
import sys


user_agents = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.79 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.53 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:101.0) Gecko/20100101 Firefox/101.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36',
]


def get_data(page, start_url):
    headers = {
        'user-agent': random.choice(user_agents),
    }
    url = start_url % page
    print(f'正在访问 {url}')
    for j in range(5):
        try:
            resp = requests.get(url, headers=headers)
            # print(resp.text)
            # with open('amzn.html', 'w', encoding='utf-8') as f:
                # f.write(resp.text)
            print(f'第{page}页状态码是:{resp.status_code},请求成功!')
            if resp.status_code == 200:
                break
        except Exception as e:
            print(f'{url} 出错了:', e)
            print(f'正在第{j+1}次重试...')
            if j == 4:
                print(f'重试了{j+1}次,请求失败,自动终止程序!')
                resp = None

    return resp


def parse_data(res):
    # 使用正则获取reviews
    pattern = r',"reviews":(.*?)},"fe'
    reviews = re.findall(pattern, res.text)[0]
    reviews = json.loads(reviews)

    return reviews


# 获取数据,并保存
def get_and_save_data(start_url, company_name, reviews_page, get_page):
    page_num = 0
    # 判断存放表格的文件夹是否存在
    if not os.path.exists('excel'):
        os.mkdir('excel')

    # 判断存放数据的excel是否存在
    if not os.path.exists(f'excel/{company_name}.xlsx'):
        # 不存在,新建
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'date'
        sheet['B1'] = 'employee_status'
        sheet['C1'] = 'review_title'
        sheet['D1'] = 'helpful'
        sheet['E1'] = 'pros'
        sheet['F1'] = 'cons'
        sheet['G1'] = 'rating_overall'
        sheet['H1'] = 'rating_balance'
        sheet['I1'] = 'rating_culture'
        sheet['J1'] = 'rating_diversity'
        sheet['K1'] = 'rating_career'
        sheet['L1'] = 'rating_comp'
        sheet['M1'] = 'rating_mgmt'
        sheet['N1'] = 'recommends'
        sheet['O1'] = 'positive_outlook'
        sheet['P1'] = 'approves_of_CEO'

        # 起始页面
        start_page = 1
        i = 2
    else:
        # 读取已有的文件
        wb = openpyxl.load_workbook(f'excel/{company_name}.xlsx')
        sheet = wb.active

        # 获取excel中最大行数
        row = sheet.max_row - 1
        start_page = row // 10 + 1
        i = (start_page - 1) * 10 + 2

    try:
        # 循环每一页拿到每一页的数据
        if reviews_page % 10 == 0:
            reviews_page = reviews_page / 10
        else:
            reviews_page = (reviews_page // 10) + 1

        if get_page == 0:
            end_page = reviews_page + 1
            print('此次要获取所有页数据!')
        else:
            end_page = start_page + get_page
            print(f'此次要获取{get_page}页数据!')

        end_page = int(end_page)

        start_page = 49
        end_page = 50

        for page in range(start_page, end_page):
            print(1, page)
            page_num = page
            print(2, page_num)
            print(f'正在访问第{page}页...')
            # 获取网页源代码
            res = get_data(page, start_url)

            # 判断是否请求成功,res为None表示未成功
            if res:
                # 解析数据
                reviews = parse_data(res)

                # 从源码中获取数据
                tree = etree.HTML(res.text)
                li_list = tree.xpath('//div[@id="ReviewsFeed"]/ol/li')

                # 循环取出每一个review,获取需要的数据
                for index, review in enumerate(reviews):
                    # 将数据写入到excel
                    date = review.get('reviewDateTime')
                    date_list = date.split('T')[0].split('-')
                    date = date_list[1] + '/' + date_list[2] + '/' + date_list[0]
                    # print(date)

                    employee_status = li_list[index].xpath('./div/div/div[1]/div[1]/span/text()')[0]
                    employee_status = employee_status.split(',')[0]
                    pros = review.get('pros').replace('', '')
                    # print('pros：', pros)
                    sheet[f'A{i}'] = date
                    sheet[f'B{i}'] = employee_status
                    sheet[f'C{i}'] = review.get('summary')
                    sheet[f'D{i}'] = review.get('countHelpful')
                    sheet[f'E{i}'] = pros
                    sheet[f'F{i}'] = review.get('cons')
                    sheet[f'G{i}'] = review.get('ratingOverall')
                    sheet[f'H{i}'] = review.get('ratingWorkLifeBalance')
                    sheet[f'I{i}'] = review.get('ratingCultureAndValues')
                    sheet[f'J{i}'] = review.get('ratingDiversityAndInclusion')
                    sheet[f'K{i}'] = review.get('ratingCareerOpportunities')
                    sheet[f'L{i}'] = review.get('ratingCompensationAndBenefits')
                    sheet[f'M{i}'] = review.get('ratingSeniorLeadership')

                    recommends = review.get('ratingRecommendToFriend')
                    positive_outlook = review.get('ratingBusinessOutlook')
                    approves_of_CEO = review.get('ratingCeo')

                    if not recommends:
                        recommends = 0
                    sheet[f'N{i}'] = recommends
                    if not positive_outlook:
                        positive_outlook = 0
                    sheet[f'O{i}'] = positive_outlook
                    if not approves_of_CEO:
                        approves_of_CEO = 0

                    # print(recommends)
                    # print(positive_outlook)
                    # print(approves_of_CEO)

                    sheet[f'P{i}'] = approves_of_CEO
                    i += 1

                    # 随机睡眠3-8秒
                    time.sleep(random.choice([k for k in range(3, 6)]))

            # 请求失败,res为空,走else,保存数据,终止程序
            else:
                print('访问网络失败,保存数据,终止程序!')
                wb.save(f'excel/{company_name}.xlsx')
                sys.exit()


    except Exception as e:
        print(f'发生错误:', e)
        print('数据已保存,稍后在进行访问!')

    print('所有请求完成,保存数据!')
    wb.save(f'excel/{company_name}.xlsx')

    return page_num, end_page


def main():
    while True:
        with open('company_info.txt', 'r', encoding='utf-8') as f_r:
            lines = f_r.readlines()
            if len(lines) < 16:
                print('company_info.txt文件请求内容为空!!!')
                break
            if not lines[15]:
                print('company_info.txt文件17行没有内容!!!')
                break
            info = lines[16]
            print(info)

        start_url, company_name, reviews_page, get_page = info.split(', ')
        reviews_page = int(reviews_page)
        get_page = int(get_page)
        print('请求的公司是：', company_name)

        while True:
            print('开始程序。。。')
            page, end_page = get_and_save_data(start_url, company_name, reviews_page, get_page)
            print(page, end_page)
            if page + 1 == end_page:
                print(1)
                print('所有数据获取完毕，终止程序！！！')
                with open('company_info.txt', 'w', encoding='utf-8') as f_w:
                    for line in lines:
                        if info in line:
                            continue
                        f_w.write(line)
                break

            print('程序中断，休息10分钟后，再继续。。。')
            time.sleep(60 * 10)


if __name__ == '__main__':
    main()
