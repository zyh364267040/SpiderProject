import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import os
import re
import json
import openpyxl
import random
from lxml import etree
import sys

from settings import *


opt = Options()
# opt.add_argument('--headless')
# opt.add_argument('--no-sandbox')
# opt.add_argument('--disable-dev-shm-usage')

driver = webdriver.Chrome(executable_path=WEBDRIVER_PATH, options=opt)


def login(url):
    # 先访问网页
    driver.get(url)
    time.sleep(1)

    # 点击 Sign In
    driver.find_element(By.XPATH, '/html/body/header/nav[1]/div/div/div/div[5]/button').click()
    time.sleep(1)

    # 输入用户名
    driver.find_element(By.NAME, 'username').send_keys(USERNAME)
    time.sleep(1)

    # 点击 Continue
    driver.find_element(By.NAME, 'submit').click()
    time.sleep(1)

    # 输入密码
    while True:
        try:
            driver.find_element(By.NAME, 'password').send_keys(PASSWORD)
        except:
            print('输入密码错误,正在重试...')
            time.sleep(2)
        else:
            break

    time.sleep(1)

    # 点击 Sign In 登录
    driver.find_element(By.NAME, 'submit').click()
    time.sleep(1)
    print('登录成功!!!')


def get_data(url):
    driver.get(url)
    return driver.page_source


def parse_data(res):
    # 使用正则获取reviews
    pattern = r',"reviews":(.*?)},"pa'
    reviews = re.findall(pattern, res)[0]
    reviews = json.loads(reviews)

    return reviews


# 获取数据,并保存
def get_and_save_data():
    # 判断存放表格的文件夹是否存在
    if not os.path.exists('excel'):
        os.mkdir('excel')

    # 判断存放数据的excel是否存在
    if not os.path.exists(f'excel/{COMPANY_NAME}.xlsx'):
        # 不存在,新建
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet['A1'] = 'date'
        sheet['B1'] = 'employee_status'
        sheet['C1'] = 'review_title'
        sheet['D1'] = 'helpful'
        sheet['E1'] = 'pros'
        sheet['F1'] = 'cons'
        sheet['G1'] = 'rating_overall'
        sheet['H1'] = 'rating_balance'
        sheet['I1'] = 'rating_culture'
        sheet['J1'] = 'rating_diversity'
        sheet['K1'] = 'rating_career'
        sheet['L1'] = 'rating_comp'
        sheet['M1'] = 'rating_mgmt'
        sheet['N1'] = 'recommends'
        sheet['O1'] = 'positive_outlook'
        sheet['P1'] = 'approves_of_CEO'

        # 起始页面
        start_page = 1
        i = 2
    else:
        # 读取已有的文件
        wb = openpyxl.load_workbook(f'excel/{COMPANY_NAME}.xlsx')
        sheet = wb.active

        # 获取excel中最大行数
        row = sheet.max_row-1
        start_page = row // 10 + 1
        i = (start_page - 1) * 10 + 2

    # try:
    # 循环每一页拿到每一页的数据
    if REVIEWS_PAGE % 10 == 0:
        reviews_page = REVIEWS_PAGE / 10
    else:
        reviews_page = (REVIEWS_PAGE // 10) + 1

    if GET_PAGE == 0:
        end_page = reviews_page + 1
        print('此次要获取所有页数据!')
    else:
        end_page = start_page + GET_PAGE
        print(f'此次要获取{GET_PAGE}页数据!')

    for page in range(start_page, end_page):
        # 拼接每页的url
        url = START_URL % page

        # 登录
        # print('开始登录...')
        # if page == start_page:
        #     try:
        #         login(url)
        #     except Exception as e:
        #         driver.get_screenshot_as_file('1.png')
        #         print('登录失败...')
        #         print(e)

        print(f'正在访问第{page}页...')
        print(url)
        # 获取网页源代码
        res = get_data(url)
        print(res)
        # driver.get_screenshot_as_file(f'{page}.png')

        # 判断是否请求成功,res为None表示未成功
        if res:
            # 解析数据
            reviews = parse_data(res)

            # 从源码中获取数据
            tree = etree.HTML(res)
            li_list = tree.xpath('//div[@id="ReviewsFeed"]/ol/li')

            # 循环取出每一个review,获取需要的数据
            for index, review in enumerate(reviews):
                # 将数据写入到excel
                date = review.get('reviewDateTime')
                date_list = date.split('T')[0].split('-')
                date = date_list[1] + '/' + date_list[2] + '/' + date_list[0]
                # print(date)

                employee_status = li_list[index].xpath('./div/div/div[1]/div[1]/span/text()')
                if employee_status:
                    employee_status = employee_status[0].split(',')[0]
                else:
                    employee_status = ''

                sheet[f'A{i}'] = date
                sheet[f'B{i}'] = employee_status
                sheet[f'C{i}'] = review.get('summary')
                sheet[f'D{i}'] = review.get('countHelpful')
                sheet[f'E{i}'] = review.get('pros')
                sheet[f'F{i}'] = review.get('cons')
                sheet[f'G{i}'] = review.get('ratingOverall')
                sheet[f'H{i}'] = review.get('ratingWorkLifeBalance')
                sheet[f'I{i}'] = review.get('ratingCultureAndValues')
                sheet[f'J{i}'] = review.get('ratingDiversityAndInclusion')
                sheet[f'K{i}'] = review.get('ratingCareerOpportunities')
                sheet[f'L{i}'] = review.get('ratingCompensationAndBenefits')
                sheet[f'M{i}'] = review.get('ratingSeniorLeadership')

                recommends = review.get('ratingRecommendToFriend')
                positive_outlook = review.get('ratingBusinessOutlook')
                approves_of_CEO = review.get('ratingCeo')

                if not recommends:
                    recommends = 0
                sheet[f'N{i}'] = recommends
                if not positive_outlook:
                    positive_outlook = 0
                sheet[f'O{i}'] = positive_outlook
                if not approves_of_CEO:
                    approves_of_CEO = 0

                sheet[f'P{i}'] = approves_of_CEO
                i += 1

            print(f'第{page}页数据保存成功!')

        # 请求失败,res为空,走else,保存数据,终止程序
        else:
            print('访问网络失败,保存数据,终止程序!')
            wb.save(f'excel/{COMPANY_NAME}.xlsx')
            sys.exit()

        # 随机睡眠3-8秒
        time.sleep(random.choice([k for k in range(3, 9)]))
    # except Exception as e:
    #     print(f'发生错误:', e)
    #     print('数据已保存,稍后在进行访问!')

    print('所有请求完成,保存数据!')
    wb.save(f'excel/{COMPANY_NAME}.xlsx')


def main():
    # 程序启动
    get_and_save_data()


if __name__ == '__main__':
    main()

