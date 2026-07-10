# -*- coding: utf-8 -*-
"""中国标准服务网 CSRES 电力行业 DL 标准信息采集脚本。

采集标准编号、标准名称、发布部门、实施日期、状态、替代情况等字段，并写入 Excel。
"""
# @Time: 2022/9/24 12:03
import sys
import requests
from lxml import etree
import openpyxl
import time
import random
from urllib.parse import urljoin
import os


user_agents = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.79 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.53 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:101.0) Gecko/20100101 Firefox/101.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A',
]


def get_data(url):
    print(url)
    headers = {
        'user-agent': random.choice(user_agents)
    }
    for i in range(5):
        try:
            res = requests.get(url, headers=headers)
            time.sleep(random.choice([2, 3, 4]))
        except Exception as e:
            print(e)
        else:
            break

    return res


def parse_and_save_data():
    # if os.path.exists('DL.xlsx'):
    #     wb = openpyxl.load_workbook('DL.xlsx')
    #     sheet = wb.active
    #     row = sheet.max_row
    #     i = row + 1
    #     start_page = row // 25 + 1
    #
    # else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet['A1'] = '标准编号'
    sheet['B1'] = '标准名称'
    sheet['C1'] = '发布部门'
    sheet['D1'] = '实施日期'
    sheet['E1'] = '状态'
    sheet['F1'] = '替代情况'
    i = 2
    start_page = 81

    for page in range(start_page, 197):
        print(f'正在获取第{page}页数据...')
        url = f'http://www.csres.com/s.jsp?keyword=DL&xx=on&wss=on&zf=on&fz=on&pageSize=25&pageNum={page}&SortIndex=1&WayIndex=0&nowUrl='
        res = get_data(url)
        print(f'第{page}页数据获取完成!!!')
        res.encoding = 'gbk'
        # print(res.text)
        tree = etree.HTML(res.text)
        trs = tree.xpath('//thead/tr')
        # print(1, trs)
        for tr in trs[1:]:
            href = tr.xpath('./td[1]/a/@href')[0]
            href = urljoin(url, href)
            # print(href)
            detail_res = get_data(href)
            if '替代情况：' in detail_res.text:
                detail_tree = etree.HTML(detail_res.text)
                detail_spans = detail_tree.xpath('//span[@class="sh14"]//text()')
                # print(detail_spans)
                flag = False
                tidai = ''
                for detail_span in detail_spans:
                    # print('detail_span', detail_span)
                    if flag:
                        if len(detail_span) == 5 and '：' in detail_span:
                            break
                        tidai += detail_span.strip().replace('\n', '').replace('\r', '')
                    if '替代情况' in detail_span:
                        flag = True
            else:
                tidai = ''

            num = tr.xpath('./td[1]//font/text()')[0]
            name = tr.xpath('./td[2]//font/text()')[0]
            part = tr.xpath('./td[3]/font/text()')[0].strip()
            date = tr.xpath('./td[4]/font/text()')
            if date:
                date = date[0]
            else:
                date = ''
            state = tr.xpath('./td[5]/font/text()')[0]
            print(num, name, part, date, state, tidai)
            # print(num, name, part, date, state)

            sheet[f'A{i}'] = num
            sheet[f'B{i}'] = name
            sheet[f'C{i}'] = part
            sheet[f'D{i}'] = date
            sheet[f'E{i}'] = state
            sheet[f'F{i}'] = tidai

            i += 1

    wb.save(f'DL{start_page}.xlsx')


def main():
    parse_and_save_data()


if __name__ == '__main__':
    main()
