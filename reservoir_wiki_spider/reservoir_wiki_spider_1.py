# -*- coding = utf-8 -*-
# @Time: 2022/10/5 14:37
import os
import time
import openpyxl
import requests
import random
from lxml import etree
from urllib.parse import urljoin


# lack_info = {
#     'name': '水库名字',
#     'zuobiao': '坐标',
#     'shuikuweizhi': '水库位置',
#     'jungongshijian': '竣工时间',
#     'shuiwei': '洪水位',
#     'weizhi': '大坝位置',
#     'shangyou': '上游流入',
#     'bagao': '最大坝高',
#     'bading': '坝顶高程',
#     'mianji': '水库面积',
#     'kurong': '最大库容',
#     'sikurong': '死库容',
#     'badingkuandu': '坝顶宽度',
#     'wangzhan': '官方网站',
#     'leixing': '大坝类型',
#     'sishuiwei': '死水位',
#     'zhengchangkurong': '正常库容',
#     'kaigongshijian': '开工时间',
#     'zongfeiyong': '总费用',
#     'badingchangdu': '坝顶长度',
#     'zhengchangshuiwei': '正常水位',
#     'guangaimianji': '灌溉面积',
#     'zuidashuishen': '最大水深',
#     'jishuimianji': '集水面积',
#     'xiayouliuchu': '下游流出',
#     'yiminshu': '移民数',
#     'zhuangjirongliang': '装机容量',
#     'jibie': '级别'
# }

lack_info = {
    '水库名字': 'name',
    '坐标': 'zuobiao',
    '水库位置': 'shuikuweizhi',
    '竣工时间': 'jungongshijian',
    '洪水位': 'shuiwei',
    '大坝位置': 'weizhi',
    '上游流入': 'shangyou',
    '最大坝高': 'bagao',
    '坝顶高程': 'bading',
    '水库面积': 'mianji',
    '最大库容': 'kurong',
    '死库容': 'sikurong',
    '坝顶宽度': 'badingkuandu',
    '官方网站': 'wangzhan',
    '大坝类型': 'leixing',
    '死水位': 'sishuiwei',
    '正常库容': 'zhengchangkurong',
    '开工时间': 'kaigongshijian',
    '总费用': 'zongfeiyong',
    '坝顶长度': 'badingchangdu',
    '正常水位': 'zhengchangshuiwei',
    '灌溉面积': 'guangaimianji',
    '最大水深': 'zuidashuishen',
    '集水面积': 'jishuimianji',
    '下游流出': 'xiayouliuchu',
    '移民数': 'yiminshu',
    '装机容量': 'zhuangjirongliang',
    '级别': 'jibie',
    '平均水深': 'pingjunshuishen'
}



user_agents = [
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.79 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.53 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:101.0) Gecko/20100101 Firefox/101.0',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.75.14 (KHTML, like Gecko) Version/7.0.3 Safari/7046A194A',
]


def get_data(url):
    headers = {
        'user-agent': random.choice(user_agents)
    }
    for i in range(5):
        try:
            print(f'正在请求{url}')
            res = requests.get(url, headers)
            time.sleep(random.uniform(0.5, 1) * 5)
            return res
        except Exception as e:
            print(e)
            print(f'请求错误,正在进行第{i+1}次重试...')

        # 将获取失败的水库写到文件中
        with open('shibai.txt', 'w', encoding='utf-8') as f:
            f.write(url.split("/")[-1] + '\n')
        print(f'{url.split("/")[-1]}获取失败...')


def parse_lack_data(res):
    # 使用xpath获取包含信息的tr
    tree = etree.HTML(res.text)
    trs = tree.xpath('//section[@class="mf-section-0"]/table/tbody/tr')
    return trs


def parse_detail_data(res):
    # 使用xpath获取详情页信息
    tree = etree.HTML(res.text)
    trs = tree.xpath('//table[@class="infobox"]/tbody/tr')
    return trs


def save_data(lack_name, lack_detail_trs, data_list):
    data_dic = {
        'name': lack_name[0]
    }
    for lack_detail_tr in lack_detail_trs:
        th_text = lack_detail_tr.xpath('./th/text()')
        if th_text:
            td_text = lack_detail_tr.xpath('./td//text()')
            # print(th_text[0], ''.join(td_text))
            try:
                data_dic[lack_info[th_text[0]]] = ''.join(td_text)
            except Exception as e:
                print(e)
            continue

        zuobiao = lack_detail_tr.xpath('./td/small/span[1]//text()')
        if zuobiao:
            for ele in zuobiao:
                if '°N' in ele and '°E' in ele:
                    zuobiao = ele
            data_dic['zuobiao'] = zuobiao
    data_list.append(data_dic)
    print(data_dic)


def data_to_excel(data_list):
    # 判断excel表格是否存在,如果不存在新建
    if os.path.exists('wiki_lack.xlsx'):
        wb = openpyxl.load_workbook('wiki_lack.xlsx')
        sheet = wb.active
        i = sheet.max_row - 1
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        i = 2

    sheet['A1'] = '水库名字'
    sheet['B1'] = '坐标'
    sheet['C1'] = '水库位置'
    sheet['D1'] = '竣工时间'
    sheet['E1'] = '洪水位'
    sheet['F1'] = '大坝位置'
    sheet['G1'] = '上游流入'
    sheet['H1'] = '最大坝高'
    sheet['I1'] = '坝顶高程'
    sheet['G1'] = '水库面积'
    sheet['K1'] = '最大库容'
    sheet['L1'] = '死库容'
    sheet['M1'] = '坝顶宽度'
    sheet['N1'] = '官方网站'
    sheet['O1'] = '大坝类型'
    sheet['P1'] = '死水位'
    sheet['Q1'] = '正常库容'
    sheet['R1'] = '开工时间'
    sheet['S1'] = '总费用'
    sheet['T1'] = '坝顶长度'
    sheet['U1'] = '正常水位'
    sheet['V1'] = '灌溉面积'
    sheet['W1'] = '最大水深'
    sheet['X1'] = '集水面积'
    sheet['Y1'] = '下游流出'
    sheet['Z1'] = '移民数'
    sheet['AA1'] = '装机容量'
    sheet['AB1'] = '级别'
    sheet['AC1'] = '平均水深'

    for data in data_list:
        sheet[f'A{i}'] = data.get('name', '')
        sheet[f'B{i}'] = data.get('zuobiao', '')
        sheet[f'C{i}'] = data.get('shuikuweizhi', '')
        sheet[f'D{i}'] = data.get('jungongshijian', '')
        sheet[f'E{i}'] = data.get('shuiwei', '')
        sheet[f'F{i}'] = data.get('weizhi', '')
        sheet[f'G{i}'] = data.get('shangyou', '')
        sheet[f'H{i}'] = data.get('bagao', '')
        sheet[f'I{i}'] = data.get('bading', '')
        sheet[f'G{i}'] = data.get('mianji', '')
        sheet[f'K{i}'] = data.get('kurong', '')
        sheet[f'L{i}'] = data.get('sikurong', '')
        sheet[f'M{i}'] = data.get('badingkuandu', '')
        sheet[f'N{i}'] = data.get('wangzhan', '')
        sheet[f'O{i}'] = data.get('leixing', '')
        sheet[f'P{i}'] = data.get('sishuiwei', '')
        sheet[f'Q{i}'] = data.get('zhengchangkurong', '')
        sheet[f'R{i}'] = data.get('kaigongshijian', '')
        sheet[f'S{i}'] = data.get('zongfeiyong', '')
        sheet[f'T{i}'] = data.get('badingchangdu', '')
        sheet[f'U{i}'] = data.get('zhengchangshuiwei', '')
        sheet[f'V{i}'] = data.get('guangaimianji', '')
        sheet[f'W{i}'] = data.get('zuidashuishen', '')
        sheet[f'X{i}'] = data.get('jishuimianji', '')
        sheet[f'Y{i}'] = data.get('xiayouliuchu', '')
        sheet[f'Z{i}'] = data.get('yiminshu', '')
        sheet[f'AA{i}'] = data.get('zhuangjirongliang', '')
        sheet[f'AB{i}'] = data.get('jibie', '')
        sheet[f'AC{i}'] = data.get('pingjunshuishen', '')
        i += 1

    wb.save('wiki.xlsx')


def main():
    # 保存数据的列表
    data_list = []
    # 所有水库页url
    lack_url = 'https://zh.m.wikipedia.org/zh-hans/%E4%B8%AD%E5%9B%BD%E5%A4%A7%E9%99%86%E5%A4%A7%E5%9E%8B%E6%B0%B4%E5%BA%93%E5%88%97%E8%A1%A8'

    # 发送请求,获取wiki所有湖信息
    res = get_data(lack_url)
    # print(res.text)

    # 解析数据
    trs = parse_lack_data(res)
    all_lack = len(trs) - 1
    for num, tr in enumerate(trs[1:]):
        # 获取每一个湖的详情url
        href = tr.xpath('./th/a/@href')
        lack_name = tr.xpath('./th/a/text()')
        if href:
            href = urljoin(lack_url, href[0])
            print(f'开始获取第{num}个湖,{lack_name}的信息,共计{all_lack}个...')
            # print(href)
            # 发送请求,获取湖详情数据
            detail_res = get_data(href)
            print(f'{lack_name}的信息获取完成...')

            # 解析详情页,获取详情页数据
            detail_trs = parse_detail_data(detail_res)
            save_data(lack_name, detail_trs, data_list)

    data_to_excel(data_list)


if __name__ == '__main__':
    main()
